"""Xuất bảng khám bệnh ra Excel.

File xuất có 2 sheet:
- "Bảng khám bệnh": đúng các cột đang hiển thị trên giao diện, mỗi ô là bản tóm
  tắt của một khối.
- "Chi tiết": mỗi trường trong template là một cột riêng — dạng bảng phẳng để
  đưa thẳng vào phân tích thống kê.

Nhãn cột và cách diễn giải giá trị đều lấy từ form_schema.json nên luôn khớp
với biểu mẫu; thêm mục mới vào template là file xuất tự có thêm cột.
"""

from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.modules.patients.form_service import FormService
from app.modules.patients.models import Examination, Patient

TITLE_FONT = Font(name="Calibri", size=14, bold=True)
META_FONT = Font(name="Calibri", size=10, italic=True, color="475569")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="0F766E")
BODY_FONT = Font(name="Calibri", size=11)
ZEBRA_FILL = PatternFill("solid", fgColor="F1F5F9")
THIN = Side(style="thin", color="CBD5E1")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

MAX_COLUMN_WIDTH = 45


def _is_leaf(node: dict[str, Any]) -> bool:
    return node.get("type") not in ("group", "block")


def _option_label(node: dict[str, Any], value: Any) -> str:
    for option in node.get("options", []):
        if option["value"] == value:
            return option["label"]
    return "" if value is None else str(value)


def format_value(node: dict[str, Any], value: Any) -> str:
    """Đổi giá trị thô thành chữ tiếng Việt như hiển thị trên giao diện."""
    if value in (None, "", [], {}):
        return ""

    node_type = node.get("type")
    if node_type == "radio":
        return _option_label(node, value)
    if node_type == "checkbox_group":
        if not isinstance(value, list):
            return str(value)
        return ", ".join(_option_label(node, v) for v in value)
    if node_type == "matrix":
        if not isinstance(value, dict):
            return ""
        parts = []
        for row in node.get("rows", []):
            cell = value.get(row["id"])
            if not cell:
                continue
            column = next((c for c in node.get("columns", []) if c["value"] == cell), None)
            parts.append(f"{row['label']}: {column['label'] if column else cell}")
        return "; ".join(parts)
    if node_type == "date":
        try:
            return date.fromisoformat(str(value)[:10]).strftime("%d/%m/%Y")
        except ValueError:
            return str(value)
    if node_type == "number" and node.get("unit"):
        return f"{value} {node['unit']}"
    return str(value)


def summarize_group(node: dict[str, Any], values: Any) -> str:
    if not isinstance(values, dict):
        return ""
    parts: list[str] = []
    for child in node.get("children", []):
        if _is_leaf(child):
            text = format_value(child, values.get(child["id"]))
            if text:
                parts.append(f"{child['label']}: {text}")
        else:
            nested = summarize_group(child, values.get(child["id"]))
            if nested:
                parts.append(nested)
    return " · ".join(parts)


def _leaf_columns(
    prefix: str, nodes: list[dict[str, Any]], path: list[str] | None = None
) -> list[tuple[str, list[str], dict]]:
    """(nhãn hiển thị, đường dẫn tới giá trị, node) cho từng cột của sheet chi tiết.

    Nhãn và đường dẫn phải cộng dồn song song: giá trị nằm sâu theo đúng cây
    group, ví dụ ``mental_exam.perception.description``.
    """
    path = path or []
    columns: list[tuple[str, list[str], dict]] = []
    for node in nodes:
        label = f"{prefix} — {node['label']}" if prefix else node["label"]
        node_path = [*path, node["id"]]
        if not _is_leaf(node):
            columns.extend(_leaf_columns(label, node.get("children", []), node_path))
        elif node["type"] == "matrix":
            # Mỗi dòng của bảng Có/Không thành một cột riêng để lọc/thống kê được.
            for row in node.get("rows", []):
                columns.append((f"{label} — {row['label']}", [*node_path, row["id"]], node))
        else:
            columns.append((label, node_path, node))
    return columns


def _dig(data: Any, path: list[str]) -> Any:
    cursor = data
    for key in path:
        if not isinstance(cursor, dict):
            return None
        cursor = cursor.get(key)
    return cursor


def _matrix_cell_label(node: dict[str, Any], value: Any) -> str:
    column = next((c for c in node.get("columns", []) if c["value"] == value), None)
    return column["label"] if column else ("" if value is None else str(value))


def _write_header(ws, headers: list[str], row: int) -> None:
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 30


def _autosize(ws, column_count: int, last_row: int) -> None:
    for col in range(1, column_count + 1):
        longest = 0
        for row in range(1, last_row + 1):
            value = ws.cell(row=row, column=col).value
            if value is None:
                continue
            longest = max(longest, max(len(part) for part in str(value).split("\n")))
        ws.column_dimensions[get_column_letter(col)].width = min(longest + 3, MAX_COLUMN_WIDTH)


class ExaminationExportService:
    def __init__(self, form_service: FormService | None = None):
        self.form_service = form_service or FormService()

    def build_workbook(
        self, patient: Patient, exams: list[Examination], exported_by: str
    ) -> BytesIO:
        disease_code = self.form_service.resolve_disease_code(
            patient.diagnosis, patient.disease_type
        )
        schema = self.form_service.get_schema(disease_code)
        block = next((b for b in schema["blocks"] if b["id"] == "examination"), None)
        groups = {child["id"]: child for child in (block or {}).get("children", [])}

        wb = Workbook()
        self._write_summary_sheet(
            wb.active, patient, exams, schema, groups, disease_code, exported_by
        )
        self._write_detail_sheet(wb.create_sheet("Chi tiết"), exams, block)
        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)
        return stream

    def _write_summary_sheet(
        self, ws, patient, exams, schema, groups, disease_code, exported_by
    ) -> None:
        ws.title = "Bảng khám bệnh"
        columns = schema.get("exam_table_columns", [])

        ws["A1"] = f"BẢNG KHÁM BỆNH — {patient.full_name} ({patient.patient_code})"
        ws["A1"].font = TITLE_FONT
        ws["A2"] = (
            f"Chẩn đoán: {patient.diagnosis or '—'}   •   "
            f"Biểu mẫu: {schema['disease_label']} ({disease_code})"
        )
        ws["A2"].font = META_FONT
        ws["A3"] = (
            f"Ngày xuất: {datetime.now().strftime('%d/%m/%Y %H:%M')}   •   "
            f"Người xuất: {exported_by}   •   Tổng số: {len(exams)} lần khám"
        )
        ws["A3"].font = META_FONT

        header_row = 5
        _write_header(ws, [c["label"] for c in columns], header_row)

        for index, exam in enumerate(exams, 1):
            data = exam.data or {}
            for col, column in enumerate(columns, 1):
                if column["key"] == "exam_info.exam_date":
                    value = exam.exam_date
                else:
                    node = groups.get(column["key"])
                    value = summarize_group(node, data.get(column["key"])) if node else ""
                cell = ws.cell(row=header_row + index, column=col, value=value or "")
                cell.font = BODY_FONT
                cell.border = BORDER
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if index % 2 == 0:
                    cell.fill = ZEBRA_FILL
                if column["key"] == "exam_info.exam_date" and value:
                    cell.number_format = "dd/mm/yyyy"
                    cell.alignment = Alignment(horizontal="center", vertical="top")

        ws.freeze_panes = ws.cell(row=header_row + 1, column=2)
        if exams:
            last = get_column_letter(len(columns))
            ws.auto_filter.ref = f"A{header_row}:{last}{header_row + len(exams)}"
        _autosize(ws, len(columns), header_row + len(exams))

    def _write_detail_sheet(self, ws, exams, block) -> None:
        columns = _leaf_columns("", (block or {}).get("children", []))
        _write_header(ws, ["STT"] + [label for label, _, _ in columns], 1)

        for index, exam in enumerate(exams, 1):
            data = exam.data or {}
            ws.cell(row=1 + index, column=1, value=index).border = BORDER
            for col, (_, path, node) in enumerate(columns, 2):
                raw = _dig(data, path)
                if node["type"] == "matrix":
                    text = _matrix_cell_label(node, raw)
                else:
                    text = format_value(node, raw)
                cell = ws.cell(row=1 + index, column=col, value=text)
                cell.font = BODY_FONT
                cell.border = BORDER
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if index % 2 == 0:
                    cell.fill = ZEBRA_FILL

        ws.freeze_panes = ws.cell(row=2, column=2)
        _autosize(ws, len(columns) + 1, 1 + len(exams))
