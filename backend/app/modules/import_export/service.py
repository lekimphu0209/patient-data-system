import csv
import io
import unicodedata
from datetime import date, datetime
from pathlib import Path


import pandas as pd
from fastapi import HTTPException, UploadFile
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy.orm import Session

from app.modules.patients.models import Patient
from app.modules.patients.repository import PatientRepository
from app.modules.patients.schemas import PatientCreate
from app.modules.patients.service import PatientService

REQUIRED_COLUMNS = ["patient_code", "full_name"]
OPTIONAL_COLUMNS = [
    "birth_date",
    "age",
    "gender",
    "hometown",
    "disease_type",
    "diagnosis",
    "status",
    "phone",
    "contact_person",
    "notes",
]

# Vietnamese labels used in error messages and in the generated template.
FIELD_LABELS = {
    "patient_code": "Mã bệnh nhân",
    "full_name": "Họ và tên",
    "birth_date": "Ngày sinh",
    "age": "Tuổi",
    "gender": "Giới tính",
    "hometown": "Quê quán",
    "disease_type": "Nhóm bệnh",
    "diagnosis": "Chẩn đoán",
    "status": "Trạng thái",
    "phone": "Số điện thoại",
    "contact_person": "Người liên hệ",
    "notes": "Ghi chú",
}

# Columns of the downloadable template, in order. Deliberately identical to the
# export sheet (minus STT) so an exported file can be edited and re-imported.
TEMPLATE_COLUMNS = [
    "patient_code",
    "full_name",
    "birth_date",
    "gender",
    "hometown",
    "phone",
    "contact_person",
    "diagnosis",
    "notes",
]

EXPORT_HEADERS = ["STT"] + [FIELD_LABELS[c] for c in TEMPLATE_COLUMNS]

DIAGNOSIS_CHOICES = ["Bình thường", "Trầm cảm", "Tâm thần phân liệt"]
GENDER_CHOICES = ["Nam", "Nữ", "Khác"]

EXPORT_DIR = Path("storage/exports")
IMPORT_DIR = Path("storage/imports")
TEMPLATE_DIR = Path("storage/templates")
TEMPLATE_FILENAME = "mau_nhap_benh_nhan.xlsx"


def _strip_accents(text: str) -> str:
    return "".join(
        ch for ch in unicodedata.normalize("NFD", text) if unicodedata.category(ch) != "Mn"
    )


def _normalize_header(value) -> str:
    """Fold a spreadsheet header into a lookup key: no accents, no case, no
    decoration such as ``*`` or a trailing ``(dd/mm/yyyy)`` hint."""
    if value is None:
        return ""
    text = str(value).strip().lower()
    if "(" in text:
        text = text.split("(", 1)[0]
    text = _strip_accents(text)
    text = text.replace("*", " ").replace("_", " ").replace("-", " ")
    return " ".join(text.split())


def _build_column_aliases() -> dict[str, str]:
    aliases: dict[str, str] = {}

    def add(field: str, *names: str) -> None:
        for name in names:
            aliases[_normalize_header(name)] = field

    # Machine-readable names (what the previous template used).
    for field in REQUIRED_COLUMNS + OPTIONAL_COLUMNS:
        add(field, field)
    # Vietnamese headers, including the ones the export sheet writes.
    add("patient_code", "Mã bệnh nhân", "Mã BN", "Mã hồ sơ")
    add("full_name", "Họ và tên", "Họ tên", "Tên bệnh nhân")
    add("birth_date", "Ngày sinh", "Ngày tháng năm sinh")
    add("age", "Tuổi")
    add("gender", "Giới tính")
    add("hometown", "Quê quán", "Địa chỉ", "Nơi ở")
    add("disease_type", "Nhóm bệnh", "Loại bệnh")
    add("diagnosis", "Chẩn đoán", "Chẩn đoán hiện tại")
    add("status", "Trạng thái")
    add("phone", "Số điện thoại", "Điện thoại", "SĐT")
    add("contact_person", "Người liên hệ", "Người thân")
    add("notes", "Ghi chú")
    return aliases


COLUMN_ALIASES = _build_column_aliases()

GENDER_TO_CODE = {
    "nam": "male",
    "nu": "female",
    "khac": "other",
    "male": "male",
    "female": "female",
    "other": "other",
}
GENDER_TO_LABEL = {"male": "Nam", "female": "Nữ", "other": "Khác"}
STATUS_TO_CODE = {
    "hoat dong": "active",
    "dang dieu tri": "active",
    "khong hoat dong": "inactive",
    "ngung dieu tri": "inactive",
    "active": "active",
    "inactive": "inactive",
}


def _is_blank(value) -> bool:
    if value is None:
        return True
    if isinstance(value, (date, datetime)):
        return False
    try:
        if pd.isna(value):
            return True
    except (TypeError, ValueError):
        return False
    return str(value).strip() == ""


def _parse_date(value):
    if _is_blank(value):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        parsed = pd.to_datetime(value, dayfirst=True)
    except Exception:
        return None
    return None if pd.isna(parsed) else parsed.date()


def _safe_int(value):
    if _is_blank(value):
        return None
    try:
        return int(float(str(value).strip()))
    except (ValueError, TypeError):
        return None


def _normalize_gender(value):
    if not value:
        return None
    return GENDER_TO_CODE.get(_normalize_header(value), value)


def _normalize_status(value):
    if not value:
        return "active"
    return STATUS_TO_CODE.get(_normalize_header(value), "active")


def _build_import_patient_data(raw: dict) -> dict:
    data = {
        "patient_code": raw.get("patient_code"),
        "full_name": raw.get("full_name"),
        "birth_date": raw.get("birth_date"),
        "age": raw.get("age"),
        "hometown": raw.get("hometown"),
        "disease_type": raw.get("disease_type"),
        "diagnosis": raw.get("diagnosis"),
        "status": _normalize_status(raw.get("status")),
    }
    contact_info = {}
    if raw.get("phone"):
        contact_info["phone"] = raw["phone"]
    if raw.get("contact_person"):
        contact_info["contact_person"] = raw["contact_person"]
    if contact_info:
        data["contact_info"] = contact_info

    patient_metadata = {}
    gender = _normalize_gender(raw.get("gender"))
    if gender:
        patient_metadata["gender"] = gender
    if raw.get("notes"):
        patient_metadata["notes"] = raw["notes"]
    if patient_metadata:
        data["patient_metadata"] = patient_metadata
    return data


def _gender_label(value: str | None) -> str:
    if not value:
        return ""
    return GENDER_TO_LABEL.get(value, value)


def _write_csv_export(patients: list[Patient], filepath: Path) -> None:
    with open(filepath, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(EXPORT_HEADERS)
        for idx, p in enumerate(patients, 1):
            metadata = p.patient_metadata or {}
            contact = p.contact_info or {}
            writer.writerow(
                [
                    idx,
                    p.patient_code,
                    p.full_name,
                    p.birth_date.strftime("%d/%m/%Y") if p.birth_date else "",
                    _gender_label(metadata.get("gender")),
                    p.hometown or "",
                    contact.get("phone", ""),
                    contact.get("contact_person", ""),
                    p.diagnosis or "",
                    metadata.get("notes", ""),
                ]
            )


# Shared look & feel for every generated workbook.
TITLE_FONT = Font(name="Calibri", size=15, bold=True, color="0F766E")
META_FONT = Font(name="Calibri", size=10, color="64748B")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", start_color="0F766E", end_color="0F766E")
BODY_FONT = Font(name="Calibri", size=11)
BOLD_FONT = Font(name="Calibri", size=11, bold=True)
HINT_FONT = Font(name="Calibri", size=10, italic=True, color="64748B")
ZEBRA_FILL = PatternFill("solid", start_color="F1F5F9", end_color="F1F5F9")
_THIN = Side(style="thin", color="CBD5E1")
THIN_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _autosize_columns(ws, column_count: int, max_row: int) -> None:
    for col in range(1, column_count + 1):
        widest = 0
        for row in range(1, max_row + 1):
            value = ws.cell(row=row, column=col).value
            if value is not None:
                widest = max(widest, len(str(value)))
        ws.column_dimensions[get_column_letter(col)].width = min(max(widest + 4, 12), 42)


def _write_xlsx_export(patients: list[Patient], filepath: Path, current_user) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Danh sách bệnh nhân"

    last_column = get_column_letter(len(EXPORT_HEADERS))
    ws.merge_cells(f"A1:{last_column}1")
    ws["A1"] = "DANH SÁCH BỆNH NHÂN"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    exported_by = (
        getattr(current_user, "full_name", None)
        or getattr(current_user, "email", None)
        or "Không xác định"
    )
    ws["A2"] = f"Ngày xuất: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = META_FONT
    ws["A3"] = f"Người xuất: {exported_by}   •   Tổng số: {len(patients)} bệnh nhân"
    ws["A3"].font = META_FONT

    header_row = 5
    for col, header in enumerate(EXPORT_HEADERS, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[header_row].height = 26

    for idx, p in enumerate(patients, 1):
        row = header_row + idx
        metadata = p.patient_metadata or {}
        contact = p.contact_info or {}
        values = [
            idx,
            p.patient_code,
            p.full_name,
            p.birth_date,
            _gender_label(metadata.get("gender")),
            p.hometown,
            contact.get("phone"),
            contact.get("contact_person"),
            p.diagnosis,
            metadata.get("notes"),
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(
                horizontal="center" if col in (1, 4, 5) else "left", vertical="center"
            )
            if idx % 2 == 0:
                cell.fill = ZEBRA_FILL
            if col == 4 and value:
                cell.number_format = "dd/mm/yyyy"

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    ws.auto_filter.ref = f"A{header_row}:{last_column}{header_row + len(patients)}"
    _autosize_columns(ws, len(EXPORT_HEADERS), header_row + len(patients))

    wb.save(filepath)


# Header captions for the template. A parenthetical hint is safe here: header
# matching strips anything from "(" onwards, so the caption stays self-documenting
# without breaking column detection.
TEMPLATE_HEADERS = {
    "patient_code": "Mã bệnh nhân *",
    "full_name": "Họ và tên *",
    "birth_date": "Ngày sinh (dd/mm/yyyy)",
    "gender": "Giới tính",
    "hometown": "Quê quán",
    "phone": "Số điện thoại",
    "contact_person": "Người liên hệ",
    "diagnosis": "Chẩn đoán",
    "notes": "Ghi chú",
}

TEMPLATE_SAMPLES = [
    ["BN001", "Nguyễn Văn An", "15/03/1985", "Nam", "Hà Nội", "0912345678",
     "Nguyễn Thị Bình", "Trầm cảm", "Tái khám định kỳ hàng tháng"],
    ["BN002", "Trần Thị Mai", "22/07/1992", "Nữ", "Nam Định", "0987654321",
     "Trần Văn Cường", "Tâm thần phân liệt", "Đang điều trị nội trú"],
    ["BN003", "Lê Hoàng Nam", "05/11/1978", "Nam", "Hải Phòng", "0901122334",
     "Lê Thị Hoa", "Bình thường", ""],
]


def _write_template_header(ws, title: str, subtitle: str) -> int:
    """Write the title block plus the header row. Returns the header row index."""
    last_column = get_column_letter(len(TEMPLATE_COLUMNS))

    ws.merge_cells(f"A1:{last_column}1")
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells(f"A2:{last_column}2")
    ws["A2"] = subtitle
    ws["A2"].font = HINT_FONT
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    header_row = 4
    for col, field in enumerate(TEMPLATE_COLUMNS, 1):
        cell = ws.cell(row=header_row, column=col, value=TEMPLATE_HEADERS[field])
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[header_row].height = 28
    return header_row


def _build_template_workbook() -> Workbook:
    """Build the import template.

    The data sheet is left empty on purpose — sample rows live on their own sheet
    so they can never be imported by accident, since the importer reads the first
    sheet only.
    """
    wb = Workbook()

    ws = wb.active
    ws.title = "Danh sách bệnh nhân"
    header_row = _write_template_header(
        ws,
        "MẪU NHẬP DANH SÁCH BỆNH NHÂN",
        "Nhập dữ liệu từ dòng ngay bên dưới. Cột có dấu (*) là bắt buộc. "
        "Xem sheet “Ví dụ mẫu” và “Hướng dẫn” để biết cách điền.",
    )

    first_data_row = header_row + 1
    last_data_row = first_data_row + 500

    # Pre-format the empty data area so typed rows keep the sheet's look.
    for row in range(first_data_row, first_data_row + 30):
        for col in range(1, len(TEMPLATE_COLUMNS) + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(
                horizontal="center" if col in (3, 4) else "left", vertical="center"
            )
            if (row - first_data_row) % 2 == 1:
                cell.fill = ZEBRA_FILL

    def add_dropdown(field: str, choices: list[str], prompt: str) -> None:
        col = get_column_letter(TEMPLATE_COLUMNS.index(field) + 1)
        validation = DataValidation(
            type="list", formula1=f'"{",".join(choices)}"', allow_blank=True
        )
        validation.promptTitle = FIELD_LABELS[field]
        validation.prompt = prompt
        validation.showInputMessage = True
        validation.errorTitle = "Giá trị không hợp lệ"
        validation.error = "Vui lòng chọn một giá trị trong danh sách."
        ws.add_data_validation(validation)
        validation.add(f"{col}{first_data_row}:{col}{last_data_row}")

    add_dropdown("gender", GENDER_CHOICES, "Chọn: " + ", ".join(GENDER_CHOICES))
    add_dropdown("diagnosis", DIAGNOSIS_CHOICES, "Chọn: " + ", ".join(DIAGNOSIS_CHOICES))

    ws.freeze_panes = ws.cell(row=first_data_row, column=1)
    for col, field in enumerate(TEMPLATE_COLUMNS, 1):
        width = max(len(TEMPLATE_HEADERS[field]), len(str(TEMPLATE_SAMPLES[0][col - 1])))
        ws.column_dimensions[get_column_letter(col)].width = min(max(width + 4, 14), 42)

    # --- Sheet 2: worked examples -------------------------------------------
    example = wb.create_sheet("Ví dụ mẫu")
    example_header = _write_template_header(
        example,
        "VÍ DỤ CÁCH ĐIỀN DỮ LIỆU",
        "Sheet này chỉ để tham khảo, hệ thống không đọc dữ liệu ở đây.",
    )
    for offset, sample in enumerate(TEMPLATE_SAMPLES):
        row = example_header + 1 + offset
        for col, value in enumerate(sample, 1):
            cell = example.cell(row=row, column=col, value=value)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(
                horizontal="center" if col in (3, 4) else "left", vertical="center"
            )
            if offset % 2 == 1:
                cell.fill = ZEBRA_FILL
    _autosize_columns(example, len(TEMPLATE_COLUMNS), example_header + len(TEMPLATE_SAMPLES))

    # --- Sheet 3: instructions ----------------------------------------------
    guide = wb.create_sheet("Hướng dẫn")
    guide["A1"] = "HƯỚNG DẪN NHẬP DỮ LIỆU"
    guide["A1"].font = TITLE_FONT
    lines = [
        "",
        "1. Nhập dữ liệu vào sheet “Danh sách bệnh nhân”, bắt đầu từ dòng ngay dưới hàng tiêu đề.",
        "2. Không đổi tên, không xóa và không đổi thứ tự các cột tiêu đề.",
        "3. Mỗi dòng tương ứng với một bệnh nhân; không để dòng trống xen giữa.",
        "",
        "Quy tắc từng cột:",
        "     • Mã bệnh nhân (*): bắt buộc, không được trùng với mã đã có trong hệ thống.",
        "     • Họ và tên (*): bắt buộc, ghi đầy đủ họ tên có dấu.",
        "     • Ngày sinh: định dạng dd/mm/yyyy, ví dụ 15/03/1985.",
        f"     • Giới tính: chọn một trong các giá trị {', '.join(GENDER_CHOICES)}.",
        f"     • Chẩn đoán: chọn một trong các giá trị {', '.join(DIAGNOSIS_CHOICES)}.",
        "     • Quê quán, Số điện thoại, Người liên hệ, Ghi chú: có thể bỏ trống.",
        "",
        "Sau khi tải file lên, hệ thống hiển thị bảng kiểm tra trước khi lưu:",
        "     • Dòng hợp lệ sẽ được ghi vào hệ thống.",
        "     • Dòng có lỗi sẽ bị bỏ qua kèm mô tả lỗi để bạn sửa lại và tải lên lần nữa.",
        "",
        "Lưu ý: file xuất dữ liệu của hệ thống dùng đúng bộ cột này, nên có thể chỉnh sửa rồi nhập lại.",
    ]
    for idx, line in enumerate(lines, start=2):
        cell = guide.cell(row=idx, column=1, value=line)
        cell.font = BOLD_FONT if line.endswith(":") else BODY_FONT
    guide.column_dimensions["A"].width = 100

    return wb


class ImportExportService:
    def __init__(self, db: Session):
        self.db = db
        self.patient_repo = PatientRepository(db)
        self.patient_service = PatientService(db)

    @staticmethod
    def _read_dataframe(file: UploadFile) -> pd.DataFrame:
        """Read an upload into a frame whose columns are internal field names.

        The header row is located by scanning rather than assumed to be first, so
        a sheet exported by this system (which carries a title block above the
        header) can be edited and re-imported unchanged.
        """
        contents = file.file.read()
        if not contents:
            raise HTTPException(status_code=400, detail="File rỗng, vui lòng chọn file khác.")

        buffer = io.BytesIO(contents)
        filename = (file.filename or "").lower()
        try:
            if filename.endswith((".xlsx", ".xlsm")):
                raw = pd.read_excel(buffer, header=None, dtype=object)
            else:
                raw = pd.read_csv(buffer, header=None, dtype=object)
        except Exception:
            raise HTTPException(
                status_code=400,
                detail="Không đọc được file. Hệ thống chỉ hỗ trợ định dạng .xlsx hoặc .csv.",
            )

        header_index = None
        for idx in range(min(len(raw), 20)):
            matched = {
                COLUMN_ALIASES[key]
                for key in (_normalize_header(v) for v in raw.iloc[idx].tolist())
                if key in COLUMN_ALIASES
            }
            if len(matched) >= 2:
                header_index = idx
                break

        if header_index is None:
            raise HTTPException(
                status_code=400,
                detail=(
                    "Không tìm thấy dòng tiêu đề hợp lệ trong file. "
                    "Vui lòng tải file mẫu và giữ nguyên tên các cột."
                ),
            )

        header_values = raw.iloc[header_index].tolist()
        df = raw.iloc[header_index + 1:].copy()
        df.columns = [
            COLUMN_ALIASES.get(_normalize_header(v), f"__ignored_{i}")
            for i, v in enumerate(header_values)
        ]
        df = df.loc[:, [not c.startswith("__ignored_") for c in df.columns]]
        df = df.loc[:, ~df.columns.duplicated()]

        data_columns = list(df.columns)
        if data_columns:
            df = df.dropna(how="all", subset=data_columns)
        # `raw` was read with header=None, so index i is spreadsheet row i + 1.
        df["__row_number"] = df.index + 1
        return df

    def preview_import_patients(self, file: UploadFile) -> dict:
        IMPORT_DIR.mkdir(parents=True, exist_ok=True)
        df = self._read_dataframe(file)

        missing_required = [c for c in REQUIRED_COLUMNS if c not in df.columns]
        if missing_required:
            labels = ", ".join(FIELD_LABELS[c] for c in missing_required)
            raise HTTPException(status_code=400, detail=f"File thiếu cột bắt buộc: {labels}.")

        rows = []
        valid_count = 0
        invalid_count = 0
        seen_patient_codes: set[str] = set()

        for _, row in df.iterrows():
            errors: list[str] = []
            raw: dict = {}

            for col in REQUIRED_COLUMNS + OPTIONAL_COLUMNS:
                if col not in df.columns:
                    continue
                value = row.get(col)
                if col == "birth_date":
                    raw[col] = _parse_date(value)
                    if raw[col] is None and not _is_blank(value):
                        errors.append(f"Ngày sinh không hợp lệ: {value}")
                elif col == "age":
                    raw[col] = _safe_int(value)
                    if raw[col] is None and not _is_blank(value):
                        errors.append(f"Tuổi không hợp lệ: {value}")
                else:
                    raw[col] = None if _is_blank(value) else str(value).strip()

            for col in REQUIRED_COLUMNS:
                if not raw.get(col):
                    errors.append(f"Thiếu thông tin bắt buộc: {FIELD_LABELS[col]}")

            patient_code = raw.get("patient_code")
            if patient_code:
                if patient_code in seen_patient_codes:
                    errors.append(f"Mã bệnh nhân “{patient_code}” bị trùng trong file")
                elif self.patient_repo.get_by_patient_code(patient_code):
                    errors.append(f"Mã bệnh nhân “{patient_code}” đã tồn tại trong hệ thống")
                else:
                    seen_patient_codes.add(patient_code)

            entry = {
                "row": int(row["__row_number"]),
                "data": _build_import_patient_data(raw),
                "valid": len(errors) == 0,
                "errors": errors,
            }
            if entry["valid"]:
                valid_count += 1
            else:
                invalid_count += 1
            rows.append(entry)

        return {
            "valid_count": valid_count,
            "invalid_count": invalid_count,
            "rows": rows,
        }

    def commit_import_patients(self, rows: list[dict]) -> dict:
        created = 0
        errors = []
        for row in rows:
            if not row.get("valid"):
                continue
            try:
                data = row["data"]
                self.patient_service.create_patient(PatientCreate(**data))
                created += 1
            except Exception as e:
                errors.append({"row": row.get("row"), "error": str(e)})
        return {"created": created, "errors": errors}

    def export_patients(
        self,
        *,
        current_user,
        ids: list[int] | None = None,
        format: str = "xlsx",
        q: str | None = None,
        diagnosis: str | None = None,
        disease_type: str | None = None,
        birth_date_from: date | None = None,
        birth_date_to: date | None = None,
    ) -> str:
        EXPORT_DIR.mkdir(parents=True, exist_ok=True)

        if ids:
            patients = (
                self.db.query(Patient)
                .filter(Patient.id.in_(ids), Patient.deleted_at.is_(None))
                .order_by(Patient.created_at.desc(), Patient.id.desc())
                .all()
            )
        else:
            patients = self.patient_repo.get_all(
                search=q,
                diagnosis=diagnosis,
                disease_type=disease_type,
                birth_date_from=birth_date_from,
                birth_date_to=birth_date_to,
            )

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        extension = "csv" if format == "csv" else "xlsx"
        filepath = EXPORT_DIR / f"danh_sach_benh_nhan_{timestamp}.{extension}"

        if extension == "csv":
            _write_csv_export(patients, filepath)
        else:
            _write_xlsx_export(patients, filepath, current_user)

        return str(filepath)

    def get_template(self) -> str:
        TEMPLATE_DIR.mkdir(parents=True, exist_ok=True)
        filepath = TEMPLATE_DIR / TEMPLATE_FILENAME
        _build_template_workbook().save(filepath)
        return str(filepath)
